import sqlite3 as sql
from yutility import log, ensure_list, dictfunc, plotfunc, listfunc, edit_distance
import os
import numpy as np
import matplotlib.pyplot as plt


python_to_sql_types = {
    str: 'TEXT',
    int: 'INTEGER',
    float: 'REAL',
    bool: 'BOOL',
    None: 'NULL',
    type: 'BLOB'
}

sql_to_python_types = {
    'TEXT': str,
    'INTEGER': int,
    'REAL': float,
    'NULL': None,
    'BOOL': bool
}


class DBSelectResult:
    def __init__(self, data, columns, types, db_path):
        self.data = data
        self.columns = columns
        self.types = types
        self.db_path = db_path

    def __str__(self):
        ret = f'DBSelectResult({len(self.data)}x{len(self.columns)}, [{", ".join(self.columns)}])'
        return ret

    def __getitem__(self, key):
        if isinstance(key, int):
            return self.data[key]
        if isinstance(key, (list, tuple, np.ndarray)):
            indices = [self.columns.index(key_) for key_ in key]
            return np.array(self.data, dtype='O')[:, indices]
        
        if key not in self.columns:
            closest = edit_distance.get_closest(key, self.columns)
            closest_dist = edit_distance.lev(closest[0], key)
            if closest_dist >=3:
                raise KeyError(f'Could not find key {key}, must be one of {", ".join(self.columns)}')
            else:
                raise KeyError(f'Could not find key {key}, did you mean {", ".join(closest)}?')
        col_idx = self.columns.index(key)
        return np.array([datum[col_idx] for datum in self.data])

    def __getattr__(self, key):
        return self.__getitem__(key)

    def __setitem__(self, key, values):
        if isinstance(key, int):
            # assert len(self.data[key]) == len(values), f'Cannot set values with length {len(self.data[key])} at index {key} with length {len(values)}'
            self.data[key] = values
        
        col_idx = self.columns.index(key)
        # assert len(self.data[key]) == len(values), f'Cannot set values with length {len(self.data[key])} at column {key} with length {len(values)}'
        newdata = []
        for val, row in zip(values, self.data):
            row = list(row)
            row[col_idx] = val
            newdata.append(row)
        self.data = newdata
            
    def groupby(self, mask_key: str):
        mask_data = self[mask_key]
        unqs = set(mask_data)
        dbs = {}
        for unq in unqs:
            dbs[unq] = DBSelectResult([x for i, x in enumerate(self) if mask_data[i] == unq], self.columns, self.types, self.db_path)
        return dbs

    def uniques(self, key: str):
        return set(self[key])

    def where(self, *masks):
        masks = ensure_list(masks)

        return DBSelectResult([x for i, x in enumerate(self) if all(mask[i] for mask in masks)], self.columns, self.types, self.db_path)

    def sortby(self, key, reference=None, sortfunc=None):
        sortval = []
        if sortfunc is None:
            sortfunc = lambda x: x
            if reference is not None:
                sortfunc = lambda x: reference.index(x)

        x = self[key]
        idx = listfunc.argsort(x, sortfunc)
        return DBSelectResult([self.data[i] for i in idx], self.columns, self.types, self.db_path)

    def remove_empty(self, keys='*', mode='all'):
        if keys == '*':
            keys = self.columns
        keys = ensure_list(keys)

        key_idxs = [self.columns.index(key) for key in keys]
        if mode == 'all':
            return DBSelectResult([x for x in self if all(x[kidx] is not None for kidx in key_idxs)], self.columns, self.types, self.db_path)
        if mode == 'any':
            return DBSelectResult([x for x in self if any(x[kidx] is not None for kidx in key_idxs)], self.columns, self.types, self.db_path)

    def __iter__(self):
        return iter(self.data)

    def iter_dict(self):
        return iter([{key: self[key][i] for key in self.columns} for i in range(len(self))])

    def __add__(self, other):
        assert isinstance(other, DBSelectResult), f'Added object must be of type DBSelectResult, not {type(other)}'
        self.data.extend(other.data)
        return self

    def __radd__(self, other):
        assert isinstance(other, DBSelectResult), f'Added object must be of type DBSelectResult, not {type(other)}'
        self.data.extend(other.data)
        return self

    def __len__(self):
        return len(self.data)

    def remove_column(self, keys):
        keys = ensure_list(keys)
        key_idxs = [self.columns.index(key) for key in keys]
        return DBSelectResult([x for i, x in enumerate(self) if all(x[kidx] is not None for kidx in key_idxs)], self.columns, self.types, self.db_path)

    def pair_plot(self, keys=None, groupkey=None, **kwargs):
        if keys is None:
            keys = self.numeric_columns
        if groupkey:
            groups = self[groupkey]
        else:
            groups = None
        return plotfunc.pair_plot([self[key] for key in keys], keys, groups=groups, groupsname=groupkey, **kwargs)

    def plot(self, xkey, ykey, groupkey=None, figsize=None, **kwargs):
        plt.figure(figsize=figsize)
        if groupkey:
            groups = self[groupkey]
        else:
            groups = None

        ykey = ensure_list(ykey)
        for key in ykey:
            shower = plotfunc.plot(self[xkey], self[key], xlabel=xkey, ylabel=', '.join(ykey), groups=groups, groupsname=groupkey, **kwargs)
        return shower

    def heatmap(self, xkey, ykey, groupkey=None, resolution=(200, 200), s2=.002, figsize=None, **kwargs):
        x, y = self[xkey], self[ykey]

        if groupkey:
            groups = self[groupkey]
            group_labels = np.unique(groups)
            Ms = [np.zeros(resolution) for _ in group_labels]
            group_indices = [np.where(groups == group_label) for group_label in group_labels]
        else:
            groups = None
            Ms = [np.zeros(resolution)]
            group_indices = [np.arange(len(x))]

        Ms_ = []
        for M, indices in zip(Ms, group_indices):
            dx = x.max() - x.min()
            xlim = x.min() - dx*.05, x.max() + dx*.05
            dy = y.max() - y.min()
            ylim = y.min() - dy*.05, y.max() + dy*.05
            X, Y = np.meshgrid(np.linspace(*xlim, resolution[0]), np.linspace(*ylim, resolution[1]))

            kernel = lambda px, py: np.exp(-((X-px)**2/(s2*dx**2) + (Y-py)**2/(s2*dy**2)))
            for px, py in zip(x[indices], y[indices]):
                M += kernel(px, py)
            M = (M - M.min()) / (M.max() - M.min())
            Ms_.append(M)

        plt.figure(figsize=figsize)
        return plotfunc.heatmap(Ms_, extent=(x.min(), x.max(), y.min(), y.max()), xlabel=xkey, ylabel=ykey, **kwargs)

    def column_type(self, key):
        return self.types[self.columns.index(key)]

    def column_of_type(self, typs):
        typs = ensure_list(typs)
        return [col for col in self.columns if self.column_type(col) in typs]

    @property
    def numeric_columns(self):
        return self.column_of_type((int, float))

    def write_excel(self, out_file: str = None):
        '''
        Write the data stored in this DBSelectResult object to an Excel file.
        '''
        import openpyxl as xl
        try:
            from openpyxl.cell import get_column_letter
        except ImportError:
            from openpyxl.utils import get_column_letter

        out_file = out_file or self.db_path.replace('.db', '.xlsx')
        # open a new notebook
        wb = xl.Workbook()

        sheet = wb.create_sheet('Data')

        # dim_holder will be used to auto-format the columns
        dim_holder = xl.worksheet.dimensions.DimensionHolder(worksheet=sheet)

        # first write the data
        for j, data in enumerate(self.data):
            for k, x in enumerate(data):
                sheet.cell(row=j+2, column=k+1, value=x)

        # write column headers
        for i, column in enumerate(self.columns):
            col_cell = sheet.cell(row=1, column=i+1, value=column)
            col_cell.font = xl.styles.Font(b=True)
            col_cell.alignment = xl.styles.Alignment(horizontal="center", vertical="center")
            col_cell.border = xl.styles.Border(bottom=xl.styles.Side(border_style="thick"))
            dim_holder.setdefault(get_column_letter(i+1), xl.worksheet.dimensions.ColumnDimension(sheet, min=i+1, max=i+1, bestFit=True))

        # fixing the column widths
        sheet.column_dimensions = dim_holder
        sheet.freeze_panes = sheet['A2']

        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        wb.save(out_file)


class DataBase:
    def __init__(self, db_path=None):
        self.is_temp = False
        self.db_path = db_path
        if db_path is None:
            self.is_temp = True
            self.db_path = 'tmp.db'

        self.name = os.path.split(db_path)[1]
        self.connection = sql.connect(self.db_path)
        self.cursor = self.connection.cursor()

    def __str__(self):
        tables = [f'{name}({self.get_table_size(name)}x{len(self.get_column_names(name))})' for name in self.get_table_names()]
        if self.is_temp:
            return f'DataBase(is_temp=True, tables=[{", ".join(tables)}])'
        return f'DataBase(file={self.name}, tables=[{", ".join(tables)}])'

    def __enter__(self):
        return self

    def close(self):
        self.connection.commit()
        self.connection.close()
        if self.is_temp:
            os.remove(self.db_path)

    def __exit__(self, *args):
        self.close()

    def __del__(self):
        try:
            self.close()
        except:
            pass

    def execute(self, command):
        while True:
            try:
                self.cursor.execute(command)
                return
            except sql.OperationalError as e:
                msg = e.args[0]
                if msg == 'database is locked':
                    log.log('Warning: database locked, retrying ...')
                else:
                    print(command)
                    raise

    def fetchall(self):
        return self.cursor.fetchall()

    def fetchone(self):
        return self.cursor.fetchone()

    def parse_type(self, typ):
        if isinstance(typ, str):
            return typ
        if isinstance(typ, np.generic):
            typ = typ.item()
        else:
            return python_to_sql_types[typ]

    def make_table(self, table_name, columns=None, types=None, primary_key=None):
        columns = columns or ['id']
        types = types or [str] * len(columns)

        assert len(columns) == len(types), f'Lengths of columns ({len(columns)}) and types ({len(types)}) have to be the same'

        tolist = []
        for colname, coltype in zip(columns, types):
            s = f'{colname} {self.parse_type(coltype)}'
            if colname == primary_key:
                s += ' PRIMARY KEY'
            tolist.append(s)
        cols = ', '.join(tolist)
        command = f"CREATE TABLE IF NOT EXISTS {table_name} ({cols})"
        self.execute(command)

    def drop_table(self, table_name):
        command = f'DROP TABLE IF EXISTS {table_name}'
        self.execute(command)

    def insert(self, table_name, values):
        vals = ', '.join([repr(x) if x is not None else 'null' for x in values])
        command = f'INSERT INTO {table_name}\nVALUES ({vals})'
        self.execute(command)

    def insert_dict(self, table_name, inp_dict, ensure_columns=False, raise_on_unknown_columns=True):
        inp_dict = dictfunc.remove_empty(inp_dict)
        column_names = inp_dict.keys()
        values = inp_dict.values()

        # handle unknown columns. If we are not ensuring columns we can raise if there are missing columns
        unknown_columns = [column for column in column_names if column not in self.get_column_names(table_name)]
        if raise_on_unknown_columns and not ensure_columns:
            if len(unknown_columns) > 0:
                raise KeyError(f"Columns [{', '.join(unknown_columns)}] not present in table '{table_name}'. Consider calling with 'ensure_columns=True'.")

        if ensure_columns:
            for col in unknown_columns:
                vals = ensure_list(inp_dict[col])
                if isinstance(vals[0], np.generic):
                    typ = type(vals[0].item())
                else:
                    typ = type(vals[0])
                self.ensure_column(table_name, col, typ)

        column_names = [column for column in column_names if column in self.get_column_names(table_name)]
        values = [inp_dict[column] for column in column_names]
        
        cols = ', '.join([repr(x) for x in column_names])
        vals = ', '.join([repr(x) for x in values])

        if not cols or not vals:
            return

        command = f'INSERT INTO {table_name} ({cols})\nVALUES ({vals})'
        self.execute(command)

    def delete_row(self, table_name, where):
        self.execute(f'DELETE FROM {table_name} WHERE {where}')

    def get_table_size(self, table_name):
        self.execute(f'SELECT Count(*) FROM {table_name}')
        return self.fetchone()[0]

    def get_column_names(self, table_name):
        self.execute(f'PRAGMA table_info({table_name})')
        return [row[1] for row in self.fetchall()]

    def get_column_types(self, table_name):
        self.execute(f'PRAGMA table_info({table_name})')
        return [sql_to_python_types[row[2].upper()] for row in self.fetchall()]

    def get_table_names(self):
        self.execute("SELECT name FROM sqlite_master WHERE type='table';")
        return [name[0] for name in self.fetchall()]

    def add_column(self, table_name, column, typ):
        command = f'ALTER TABLE {table_name} ADD COLUMN "{column}" {self.parse_type(typ)}'
        self.execute(command)

    def delete_column(self, table_name, column):
        ...

    def ensure_column(self, table_name, column, typ):
        if column not in self.get_column_names(table_name):
            self.add_column(table_name, column, typ)

    def select(self, table_name, columns=None, where=None):
        cols = '*'
        if columns is not None:
            columns = ensure_list(columns)
            cols = ', '.join(columns)
        if cols == '*':
            columns = self.get_column_names(table_name)

        # types = [self.get_column_types(table_name)[i] for i, col in enumerate(self.get_column_names(table_name)) if col in columns]
        types = [self.get_column_types(table_name)[self.get_column_names(table_name).index(col)] for col in columns]
        if where is not None:
            where = 'WHERE ' + where
        command = f'SELECT {cols}\n\tFROM {table_name}\n\t{where}'
        self.execute(command)
        result = self.fetchall()
        return DBSelectResult(result, columns, types, self.db_path)

    def delete_duplicates(self, table_name, columns=None):
        cols = '*'
        if columns is not None:
            columns = ensure_list(columns)
            cols = ', '.join(columns)
        command = rf"""
        DELETE FROM {table_name}
        WHERE rowid NOT IN (
          SELECT 
            MIN(rowid) 
          FROM 
            {table_name} 
          GROUP BY
            {', '.join(cols)}
        )"""
        print(command)


def merge_databases(databases, new_name):
    databases = ensure_list(databases)
    if isinstance(databases[0], str):
        databases = [DataBase(db_path) for db_path in databases]

    with DataBase(new_name) as db:
        for db_old in databases:
            for table in db_old.get_table_names():
                db.make_table(table)
                cols = db_old.get_column_names(table)
                data = db_old.select(table, '*')
                for datum in data:
                    datum_dict = {col: value for col, value in zip(cols, datum)}
                    db.insert_dict(table, datum_dict, ensure_columns=True)

    return DataBase(new_name)
